import glob
import zipfile
from os.path import basename
from zipfile import ZipFile
import csv
import pytest
from PyPDF2 import PdfReader
from zipfile import ZipFile
import os

from openpyxl.reader.excel import load_workbook

dir_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
path_zip = os.path.join(dir_path, 'hw_archive.zip')


path_to_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')

path_to_csv = os.path.join(path_to_resources, 'doc2.csv')
path_to_pdf = os.path.join(path_to_resources, 'exmpl.pdf')
path_to_xlsx = os.path.join(path_to_resources, 'testdoc1.xlsx')


#Удаление файлов в zip
@pytest.fixture()
def clear_file_from_zip():
    all_files = os.path.join(path_zip, '*.*')
    for file in glob.glob(all_files):
        os.remove(file)


#Создаем zip-архив
def test_create_zip(clear_file_from_zip):
    with ZipFile(path_zip, 'w') as myzip:
        myzip.write(path_to_csv, basename(path_to_csv))
        myzip.write(path_to_pdf, basename(path_to_pdf))
        myzip.write(path_to_xlsx, basename(path_to_xlsx))


  #Чтение и проверка pdf
def test_read_pdf():
    #hw_zipfile = ZipFile(path_zip)
    with zipfile.ZipFile(path_zip) as hw_zipfile:
        r_pdf = hw_zipfile.extract('exmpl.pdf')
        reader1 = PdfReader(r_pdf)
        text = reader1.pages[0].extract_text()
        print(text)
        assert 'Пример pdf' in text
        os.remove(r_pdf)


#Чтение и проверка csv
def test_read_csv():
    hw_zipfile = ZipFile(path_zip)
    text = str(hw_zipfile.read('doc2.csv'))
    print(text)
    assert 'Elena' in text



#Чтение и проверка xlsx
def test_read_xlsx():
    hw_zipfile = ZipFile(path_zip)
    workbook = hw_zipfile.extract('testdoc1.xlsx')
    xlsx_book = load_workbook(workbook)
    sheet = xlsx_book.active
    print(sheet.cell(row=1, column=1).value)
    assert 'Desktop' in sheet.cell(row=1, column=1).value
    os.remove(workbook)





