import openpyxl as openpyxl
import pytest
import pypdf
import csv
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from zipfile import ZipFile


#Чтение и проверка pdf
def test_read_pdf():
    hw_zipfile = ZipFile('../hw_archive.zip')
    r_pdf = hw_zipfile.open('../resources/exmpl.pdf')
    reader1 = PdfReader(r_pdf)
    text = reader1.pages[0].extract_text()
    print(text)
    assert 'Пример pdf' in text


#Чтение и проверка csv
def test_read_csv():
    hw_zipfile = ZipFile('../hw_archive.zip')
    text = str(hw_zipfile.read('resources/doc2.csv'))
    print(text)
    assert 'Elena' in text


#Чтение и проверка xlsx
def test_read_xlsx():
    hw_zipfile = ZipFile('../hw_archive.zip')
    workbook = load_workbook(hw_zipfile.open('../resources/testdoc1.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=1, column=1).value)
    assert 'Desktop' in sheet.cell(row=1, column=1).value




